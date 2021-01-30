# Set paths
import os
import pandas as pd

import datetime as dt
import calendar
calendar.setfirstweekday(calendar.SUNDAY)  
today = dt.datetime.today().strftime("%Y.%m.%d")

import pickle
from openpyxl import Workbook, load_workbook

# Set paths
import os
import random

# Data manipulation
import pandas as pd
import numpy as np

# Date manipulation
import datetime as dt
import calendar
calendar.setfirstweekday(calendar.SUNDAY) 

# Geolocation
import geonamescache


def clean_date_format(df, date_list):
    for date in date_list:
        df[date] = df[date].str.split('.', n=1, expand=True)[0]
    return df

def impute_nan(df, features, nan_replacements=' '):
    for feature in features:
        df[feature] = df[feature].replace(to_replace=np.nan, value=nan_replacements)
    return df

def impute_missing_dates(df, identifier, dates, fillna_method='bfill'):
    for date in dates:
        df       = df.sort_values([identifier, date], ascending=True).reset_index(drop=True)
        df[date] = df[date].fillna(method=fillna_method)
        df       = df.sort_values([date], ascending=True).reset_index(drop=True)
    return df

def derive_datetime(df, date_list):
    for date in date_list:
        df[date]                   = pd.to_datetime(df[date])
        df[date+'.date']           = df[date].dt.date
        df[date+'.year']           = df[date].dt.year
        #df[date+'.month']          = df[date].dt.month.apply(lambda x: calendar.month_abbr[x])
        df[date+'.month_num']      = df[date].dt.month
        #df[date+'.year_month']     = df[date+'.year'].map(str) + " " + df[date+'.month'].map(str) 
        #df[date+'.year_month_num'] = df[date+'.year'].map(str) + " " + df[date+'.month_num']
        df[date+'.hour']           = df[date].dt.hour
        df[date+'.day']            =df[date].dt.day
        #df[date+'.weekday']        = df[date].dt.weekday_name
    return df

def drop_selected_columns(df, features):
    df = df.drop(features, axis=1)
    return df

def create_invoice_no(df, date, identifier):
    df['InvoiceNo'] = df.groupby([date, identifier]).ngroup()+1
    return df

def create_sequential_identifier(df, identifier):
    df[identifier+'.clean'] = np.nan
    df[identifier+'.clean'] = df.index + 1
    df.reset_index(drop=True, inplace=True)
    return df

def generate_state_given_city(df, city_list, col_name):
#     import geonamescache
    gc = geonamescache.GeonamesCache()
    state_list = []  
    
    for city in city_list:
        info = gc.get_cities_by_name(city)
        if info == []:
            state_list.append(np.nan)
        else:
            for dictionary in list(info[0].values()):
                state = list(dictionary.values())
                state_list.append(state[7])
    df[col_name] = state_list
    return df

def categorize_items(df, item_feature, new_category_feature, items_dict): 
    category = ' '
    df[new_category_feature] = category

    for k, v in items_dict.items():
        pat = '|'.join(v)
        df[item_feature] = df[item_feature].str.strip()
        df[item_feature] = df[item_feature].str.title()
        mask = df[item_feature].str.contains(pat, case=True)

        df.loc[mask, new_category_feature] = k
    return df

   
def categorize_customers(rfm_score):
    x = rfm_score
    if (x == '111'):
        return 'best customers'
    elif (x == '311'):
        return 'almost lost'
    elif (x == '411'):
        return 'lost customers'
    elif (x == '444'):
        return 'lost cheap customers'
    else:
        return 'other'        


def timer(start_time, text=""):
    '''
    Indicates run time of script.

    Parameters
    ----------
    start_time : time
        The time the script begins to run. 

    text : str
        Optional text.

    Returns
    -------
    time
        Run time of script.

    '''
    
    current_time = time.time()
    print(text + "Runtime is %s seconds" % (current_time - start_time))
    #return current_time


def read_excel_sheet(wb, sheet_name):
    sheet = pd.read_excel(wb, sheet_name=sheet_name, skiprows=1)
    sheet = sheet.loc[:, ~sheet.columns.str.contains('^Unnamed: ')]
    return sheet

def concat(df_list):
    df = pd.concat(df_list, axis=1, sort=False)
    return df

def read_concat_multiple_sheets(wb, sheets):
    df_list = []
    for sheet_name in sheets:
        df = read_excel_sheet(wb, sheet_name)
        df_list.append(df)
        df_combi = concat(df_list)
    return df_combi

# def load_workbook(wb, desired_sheets)
# #desired_sheets = ['2018 Orders', '2019 Orders']
# lagos_xlsx = pd.ExcelFile(lagos_file)
# lagos_sheets = []
# for sheet in lagos_xlsx.sheet_names:
#     if sheet in desired_sheets:
#         lagos_sheets.append(lagos_xlsx.parse(sheet))
#         lagos_sub_xlsx = pd.concat(lagos_sheets, sort=True)

# null_columns=lagos_data_2016.columns[lagos_data_2016.isnull().any()]
# lagos_data_2016[null_columns].isnull().sum()

def getDuplicateColumns(df):
    '''
    Get a list of duplicate columns.
    It will iterate over all the columns in dataframe and find the columns whose contents are duplicate.
    :param df: Dataframe object
    :return: List of columns whose contents are duplicates.
    '''
    df.columns = df.columns.str.strip()
    duplicateColumnNames = set()
    # Iterate over all the columns in dataframe
    for x in range(df.shape[1]):
        # Select column at xth index.
        col = df.iloc[:, x]
        # Iterate over all the columns in DataFrame from (x+1)th index till end
        for y in range(x + 1, df.shape[1]):
            # Select column at yth index.
            otherCol = df.iloc[:, y]
            # Check if two columns at x 7 y index are equal
            if col.equals(otherCol):
                duplicateColumnNames.add(df.columns.values[y])
 
    return list(duplicateColumnNames)

    # Get list of duplicate columns
    duplicateColumnNames = getDuplicateColumns(df)
 
    print('Duplicate Columns are as follows')
    for col in duplicateColumnNames:
        print('Column name : ', col)

       
       
def drop_dup_col_diff_label(df):
    #Delete duplicate columns
    newDf = df.drop(columns=getDuplicateColumns(df))

    return newDf

def drop_dup_col_same_label(df):
    df = df.loc[:,~df.columns.duplicated()]
    return df

def ordered_factor(_df, old_colname, new_colname, ordered_categories):
    _df[new_colname] = pd.Categorical(_df[old_colname], ordered=True, categories=ordered_categories)
    return _df

def missing_data_table(_df):
    '''
    Shows the distribution (count and proportion) of variables with missing values.
    Parameters
    ----------
    _df : dataframe
    The data set/dataframe being checked for missing data.. 
    '''
    if (_df.isnull().any().sum() > 0):
        missing = _df.isnull().sum()[_df.isnull().sum() > 0]
        missing_df = pd.DataFrame({'Variable':missing.index, 'Count':missing.values, 'Proportion': (missing.values/len(_df))})
        missing_df['Proportion'] = missing_df['Proportion'].round(decimals=2)
        print("Missing data distribution:\n\n{}".format(missing_df.to_string(index=False)))
    
        # t = doc.add_table(missing_df.shape[0]+1, missing_df.shape[1])

        # # add the header rows.
        # for j in range(missing_df.shape[-1]):
        #     t.cell(0,j).text = missing_df.columns[j]

        # # add the rest of the data frame
        # for i in range(missing_df.shape[0]):
        #     for j in range(missing_df.shape[-1]):
        #         t.cell(i+1,j).text = str(missing_df.values[i,j])
        
    else:
        print("There is no missing data")
        
#This function takes in as argument a dataframe and n-size list of columns to output the unique count
def check_unique_no(df, col_list):
    for col in col_list:
        print('Data has {} unique {}'.format(df[col].nunique(), col))


# def chron_list():
#     for i,col in enumerate(data.columns):
#     print(i+1,". column is ",col)


# def subset_data(_df):
# # Get target
# target = 'loan_status'
# print('The target variable is {}'.format(target))

# # Subset nominal categorical features and remove the unnecessary features
# categorical = _df.select_dtypes(exclude=['int64', 'float64', 'datetime64[ns]', 'category']).columns.tolist()
# cat_var_to_remove = {'next_pymnt_d'}
# categorical = [f for f in categorical if f not in cat_var_to_remove]
# print('The number of nominal categorical variables to display are {}'.format(len(categorical)))
# df_nominal = df1.filter(categorical)

# # Subset ordinal categorical features and delete the unnecessary features
# ord_categorical = df1.select_dtypes(include=['category']).columns.tolist()
# print('The number of ordinal categorical variables to display are {}'.format(len(ord_categorical)))
# df_ordinal = df1.filter(ord_categorical)

# # Subset numerical features and delete the unnecessary features
# numerical = df1.select_dtypes(exclude=['object', 'category', 'datetime64[ns]']).columns.tolist()
# num_var_to_remove = {'out_prncp', 'out_prncp_inv', 'policy_code'}
# numerical = [f for f in numerical if f not in num_var_to_remove]
# print('The number of numerical variables to display are {}'.format(len(numerical)))
# df_numerical = df1.filter(numerical)

def save_pickle(_df, pickle_path):
    f = open(pickle_path, 'wb')
    pickle.dump(_df, f)
    f.close()

def load_pickel(_df, pickle_path):
    f = open(pickle_path, 'rb')
    pickle.load(_df, f)
    f.close()

def columns_preprocess(_df):
    _df.columns = [c.lower().replace(' ', '_') for c in _df.columns]
    return _df

def label_encode(_df, cols):
    for col in cols:
        _df[col] = _df[col].astype('category')
        _df[col+'.cat'] = _df[col].cat.codes
    return _df

def customer_profile(_df, unique_id, value, featurelist):
    df_list = []
    for feature in featurelist:
        _df[feature] = _df[feature].astype(str)
        _df_temp = pd.crosstab(_df[unique_id], _df[feature], values=_df[value], 
                               aggfunc=['count','sum', 'mean', 'max', 'min', pd.Series.nunique], dropna=False).fillna(0).reset_index()
        _df_temp.columns = [".".join(x).strip('.') for x in _df_temp.columns.ravel()] 
        df_list.append(_df_temp)
        dfs = [df.set_index(unique_id) for df in df_list]
        df = pd.concat(dfs, axis=1)
        df.reset_index(inplace=True)
    return df

# def rename_csv(df_path, dict_path, sheet_name, data_set): 
#     '''
#     Rename data set variables to SFL standard names. 
#     Also indicates the number of rows and columns in the dfset as well as
#     the number of unique customer numbers in the df set.

#     Parameters
#     ----------
#     df_path : str
#         The path from which the df set should be read

#     dict_path : str
#         The path from which the dictionary should be read
    
#     sheet_name : str
#         The name of the sheet/tab of the dictionary in the excel workbook

#     data_set : str
#         The name of the df set with the variables to be renamed

#     Returns
#     -------
#     dataframe
#         A dataframe with variables in SFL standard names

#     '''
   
#     df = pd.read_csv(df_path, mangle_dupe_cols=True)
#     _dict = pd.read_excel(dict_path, sheet_name=sheet_name)
#     _dict = _dict.dropna(thresh=4)
#     _dict = _dict.loc[_dict['Data Set'] == data_set, :]
#     df.rename(columns=dict(zip(_dict['External'],_dict['SFL'])), inplace=True)
    
#     df_new = df.filter(items=_dict['SFL'].unique().tolist())
    
#     all_colname = df_new.columns.tolist()
#     all_colname = ', '.join(str(v) for v in all_colname)
    
#     print("The variables in the df are: \n{}".format(all_colname))
    
#     print("\nThe df has {:,.0f} records and {:,.0f} variables".format(len(df_new), df_new.shape[1]))
#     print("\nThe df contains {:,.0f} unique customer numbers".format(df_new['customer_no'].nunique()))

#     return df_new

# #Extract feature names in data
# def extract_feature_names(df, data_dictionary_path, dataset):
#     wb = load_workbook(data_dictionary_path)
#     wb = pd.ExcelWriter(issues_file_path, engine = 'xlsxwriter')
            
#             for sheet_name in dfs.keys():
#                 dfs[sheet_name].to_excel(wb, sheet_name=sheet_name)
#             wb.save()
#             wb.close()
#     df = pd.DataFrame({'Data Set': dataset,
#                        'Client Feature Name': list(df.columns),
#                        'Standardized Feature Name':})
#     return df

# def data_preview(df, threshold=0.20):
#     _df = pd.DataFrame(df.dtypes,columns=['data type'])
#     _df = _df.reset_index()
#     _df['feature'] = _df['index']
#     _df = _df[['feature','data type']]
#     _df['first value'] = df.loc[0].values
#     _df['missing value count'] = df.isnull().sum().values
#     _df['missing value proportion'] = df.isnull().sum().values/len(df)
#     _df['description'] = np.nan
#     _df['questions'] = np.nan
#     _df['answers'] = np.nan
#     print('The data has {} features and {} data records'.format(df.shape[1], df.shape[0]))
#     print('\n{} features have more than {} percent of data points missing. These features include: \n'.format(len(list(_df.loc[_df['missing value proportion'] > threshold]['feature'])), int(threshold * 100)))
#     features_with_missing_data = list(_df.loc[_df['missing value proportion'] > threshold]['feature'])
#     print(*features_with_missing_data, sep=', ')
#     return _df